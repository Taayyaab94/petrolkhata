"""Rendering-only module for the PDF/Excel report exports.

Deliberately has no import of app.py (or anything Flask-specific) to avoid
a circular import - routes hand this module plain data (strings, numbers,
lists of tuples) built from the same query logic the HTML page already
uses, and get back an in-memory BytesIO buffer to send_file(). Nothing
here touches the database or the request context.

Every export is described as a list of `blocks`, rendered in order by
whichever of build_pdf/build_xlsx is called - the same block list can
usually feed both, since each renderer decides for itself how to display
a value (build_pdf formats a float for reading; build_xlsx keeps it as a
real number so it's summable in Excel).

    {"type": "summary", "rows": [(label, value), ...]}
    {"type": "table", "heading": str, "columns": [...], "rows": [[...], ...], "align": ["left", "right", ...]}
    {"type": "note", "text": str}
"""

import io
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BRAND = "Petrol Khata"


def _is_number(value):
    # bool is technically an int subclass in Python - a checkbox-like True/
    # False value must never be right-aligned or number-formatted as 1/0.
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _format_cell_text(value):
    """Display text for a PDF cell - build_xlsx keeps the raw value
    instead, since real numbers (not this formatted string) are the
    point of offering Excel at all."""
    if value is None:
        return "-"
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value)


# ---------------------------------------------------------------- PDF -----

def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=16, spaceAfter=2))
    styles.add(ParagraphStyle("ReportSubtitle", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#64748b")))
    styles.add(ParagraphStyle("BlockHeading", parent=styles["Heading3"], fontSize=12, spaceBefore=10, spaceAfter=6))
    styles.add(ParagraphStyle("CellLeft", parent=styles["Normal"], fontSize=8, leading=10, alignment=TA_LEFT))
    styles.add(ParagraphStyle("CellRight", parent=styles["Normal"], fontSize=8, leading=10, alignment=TA_RIGHT))
    styles.add(ParagraphStyle("CellHeadLeft", parent=styles["Normal"], fontSize=8, leading=10, alignment=TA_LEFT, textColor=colors.HexColor("#1f2937")))
    styles.add(ParagraphStyle("CellHeadRight", parent=styles["Normal"], fontSize=8, leading=10, alignment=TA_RIGHT, textColor=colors.HexColor("#1f2937")))
    styles.add(ParagraphStyle("Note", parent=styles["Normal"], fontSize=8.5, textColor=colors.HexColor("#64748b")))
    return styles


def _pdf_summary_table(rows, styles):
    data = []
    for label, value in rows:
        data.append([
            Paragraph(str(label), styles["CellLeft"]),
            Paragraph(_format_cell_text(value), styles["CellRight"]),
        ])
    table = Table(data, colWidths=[None, 120])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#e2e8f0")),
            ]
        )
    )
    return table


def _pdf_data_table(block, styles):
    columns = block["columns"]
    align = block.get("align") or ["left"] * len(columns)
    head_styles = [styles["CellHeadRight"] if a == "right" else styles["CellHeadLeft"] for a in align]
    body_styles = [styles["CellRight"] if a == "right" else styles["CellLeft"] for a in align]

    data = [[Paragraph(str(c), s) for c, s in zip(columns, head_styles)]]
    for row in block["rows"]:
        data.append([Paragraph(_format_cell_text(v), s) for v, s in zip(row, body_styles)])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _pdf_header_footer(canvas, doc, title):
    canvas.saveState()
    width, height = A4
    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawString(18 * mm, height - 14 * mm, BRAND)
    canvas.setFont("Helvetica", 9)
    canvas.setFillColor(colors.HexColor("#475569"))
    canvas.drawRightString(width - 18 * mm, height - 14 * mm, title)
    canvas.setStrokeColor(colors.HexColor("#cbd5e1"))
    canvas.line(18 * mm, height - 16.5 * mm, width - 18 * mm, height - 16.5 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawCentredString(width / 2, 10 * mm, f"Generated {date.today().isoformat()} - page {doc.page}")
    canvas.restoreState()


def build_pdf(title, subtitle, blocks):
    """Render `blocks` (see module docstring) as an A4 PDF, returned as an
    in-memory BytesIO ready for send_file(). Table blocks repeat their
    header row across a page break (repeatRows=1) since a report can
    easily run longer than one page."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=22 * mm,
        bottomMargin=16 * mm,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        title=title,
    )
    styles = _pdf_styles()

    story = [Paragraph(title, styles["ReportTitle"])]
    if subtitle:
        story.append(Paragraph(subtitle, styles["ReportSubtitle"]))
    story.append(Spacer(1, 10))

    for block in blocks:
        if block["type"] == "summary":
            story.append(_pdf_summary_table(block["rows"], styles))
            story.append(Spacer(1, 12))
        elif block["type"] == "table":
            if block.get("heading"):
                story.append(Paragraph(block["heading"], styles["BlockHeading"]))
            if block["rows"]:
                story.append(_pdf_data_table(block, styles))
            else:
                story.append(Paragraph("No entries.", styles["Note"]))
            story.append(Spacer(1, 12))
        elif block["type"] == "note":
            story.append(Paragraph(block["text"], styles["Note"]))
            story.append(Spacer(1, 8))

    doc.build(
        story,
        onFirstPage=lambda c, d: _pdf_header_footer(c, d, title),
        onLaterPages=lambda c, d: _pdf_header_footer(c, d, title),
    )
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------- XLSX -----

_ILLEGAL_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')


def _safe_sheet_name(name, used):
    """Excel sheet names are capped at 31 chars and can't contain
    \\ / * ? : [ ] - collide-proofed with a numeric suffix so two
    same-named sheets (unlikely here, but cheap to guard) don't clobber
    each other."""
    cleaned = _ILLEGAL_SHEET_CHARS.sub("-", name).strip() or "Sheet"
    candidate = cleaned[:31]
    n = 1
    while candidate in used:
        suffix = f" ({n})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


class _ColumnWidths:
    """Tracks the widest cell seen per column so widths can be sized from
    content, capped so one long note doesn't blow out the sheet."""

    def __init__(self, cap=40, minimum=8):
        self.cap = cap
        self.minimum = minimum
        self.widths = {}

    def see(self, col, value):
        if value is None:
            return
        length = len(_format_cell_text(value)) if not isinstance(value, str) else len(value)
        current = self.widths.get(col, self.minimum)
        self.widths[col] = min(max(current, length + 2), self.cap)

    def apply(self, worksheet):
        for col, width in self.widths.items():
            worksheet.column_dimensions[get_column_letter(col)].width = width


_HEADER_FILL = PatternFill("solid", fgColor="EEF2F7")
_MONEY_FORMAT = "#,##0.00"


def build_xlsx(sheets):
    """Render `sheets` (list of {"name", "blocks"}, blocks shaped as in
    the module docstring) as an in-memory .xlsx workbook, one worksheet
    per entry. Numeric values are written as real numbers (not strings)
    with a #,##0.00 format, so totals in the exported file are actually
    summable in Excel rather than just readable."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    for sheet in sheets:
        ws = wb.create_sheet(_safe_sheet_name(sheet["name"], used_names))
        widths = _ColumnWidths()
        row = 1
        table_block_count = sum(1 for b in sheet["blocks"] if b["type"] == "table")

        for block in sheet["blocks"]:
            if block["type"] == "summary":
                for label, value in block["rows"]:
                    label_cell = ws.cell(row=row, column=1, value=label)
                    label_cell.font = Font(bold=True)
                    value_cell = ws.cell(row=row, column=2, value=value)
                    if _is_number(value):
                        value_cell.number_format = _MONEY_FORMAT
                    widths.see(1, label)
                    widths.see(2, value)
                    row += 1
                row += 1

            elif block["type"] == "table":
                if block.get("heading"):
                    heading_cell = ws.cell(row=row, column=1, value=block["heading"])
                    heading_cell.font = Font(bold=True, size=13)
                    row += 1
                header_row_index = row
                for col_idx, name in enumerate(block["columns"], start=1):
                    cell = ws.cell(row=row, column=col_idx, value=name)
                    cell.font = Font(bold=True)
                    cell.fill = _HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
                    widths.see(col_idx, name)
                row += 1
                for data_row in block["rows"]:
                    for col_idx, value in enumerate(data_row, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=value)
                        if _is_number(value):
                            cell.number_format = _MONEY_FORMAT
                        widths.see(col_idx, value)
                    row += 1
                # Only freeze when this sheet has exactly one table - with
                # several tables stacked on a sheet there's no single
                # header row that unambiguously deserves to stay pinned.
                if table_block_count == 1:
                    ws.freeze_panes = ws.cell(row=header_row_index + 1, column=1).coordinate
                row += 1

            elif block["type"] == "note":
                note_cell = ws.cell(row=row, column=1, value=block["text"])
                note_cell.font = Font(italic=True, size=9)
                widths.see(1, block["text"])
                row += 2

        widths.apply(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
